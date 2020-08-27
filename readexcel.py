import openpyxl

book = openpyxl.load_workbook(r'E:/新建 XLSX 工作表.xlsx')
sheet = book.worksheets[2]

list = []
rows = sheet.max_row
cols = 2
for i in range(2,rows+1):
    content = sheet.cell(i,cols).value
    list.append(content)
print (i)
print(list)